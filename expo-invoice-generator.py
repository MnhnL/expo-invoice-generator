import csv
import os.path
from collections import defaultdict
import datetime
import platform

from openpyxl import load_workbook

from fpdf import FPDF, XPos, YPos, Align
from fpdf.enums import MethodReturnValue, WrapMode

FONT_SZ_HEADER = 14
FONT_SZ_FOOTER = 7

FONT_SZ_TABLE_HEADER = 10
FONT_SZ_TABLE_ROW = 8

FONT_SZ_NORMAL = 10

if platform.system() == 'Linux':
    FONT_PATH = "/usr/share/fonts/truetype/dejavu"
    FONT_NAME = "DejaVuSans"
    FONT_FILE_REGULAR = "DejaVuSans.ttf"
    FONT_FILE_OBLIQUE = "DejaVuSans-Oblique.ttf"
    FONT_FILE_BOLD = "DejaVuSans-Bold.ttf"
elif platform.system() == 'Windows':
    FONT_PATH = "C:\\Windows\\Fonts"
    FONT_NAME = "Segoe UI"
    FONT_FILE_REGULAR = "segoeui.ttf"
    FONT_FILE_OBLIQUE = "segoeuii.ttf"
    FONT_FILE_BOLD = "segoeuib.ttf"

ROW_HEIGHT_HEADER = 4
ROW_HEIGHT_ROW = 5

def format_price(p):
    return f"{p:,.2f} €".translate(str.maketrans(",.", ".,"))

def format_booking_number(bn_int):
    bn = str(bn_int)
    return f"EX-{bn[0:3]}-{bn[3:7]}"

# Initialize FPDF class
class PDF(FPDF):
    widths = (30, 40, 35, 47, 25)
    local_page_no = 0

    def new_document(self):
        self.local_page_no = self.page_no()
    
    def header(self):
        self.set_font(FONT_NAME, "B", FONT_SZ_HEADER)
        self.cell(0, 10, "Relevé des visites organisées par le MNHN", align=Align.C, border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        self.ln(8)

    def footer(self):
        # Go to 1.5 cm from bottom
        self.set_y(-15)
        self.set_font(FONT_NAME, "", FONT_SZ_FOOTER)
        self.cell(0, 10, f"Page {self.page_no() - self.local_page_no + 1}", align=Align.R, new_x=XPos.RIGHT, new_y=YPos.TOP)

    def row(self, height, cells, border="B",
            styles=("", "", "", ""),
            aligns=("L", "L", "L", "L")):

        # Measure cell height in lines
        cell_col_lines = []
        for i, el in enumerate(zip(cells, self.widths, aligns, styles)):
            text, width, a, style = el
            if i < len(cells)-1:
                new_x = XPos.RIGHT
                new_y = YPos.TOP
            else: # last col
                new_x = XPos.LMARGIN
                new_y = YPos.NEXT

            self.set_font(FONT_NAME, style, FONT_SZ_TABLE_ROW)
            cell_lines = self.multi_cell(width, height, text, border=border, align=a,
                                         new_x=new_x, new_y=new_y, wrapmode=WrapMode.CHAR,
                                         dry_run=True, output=MethodReturnValue.LINES)
            cell_col_lines.append(cell_lines)

        max_lines = max([len(cell_lines) for cell_lines in cell_col_lines])

        # Actually output cells
        for i, el in enumerate(zip(cell_col_lines, self.widths, aligns, styles)):
            cell_lines, width, align, style = el
            if i < len(cells)-1:
                new_x = XPos.RIGHT
                new_y = YPos.TOP
            else: # last col
                new_x = XPos.LMARGIN
                new_y = YPos.NEXT

            self.set_font(FONT_NAME, style, FONT_SZ_TABLE_ROW)
            empty_lines = max_lines - len(cell_lines)
            text = "\n".join(cell_lines) + "".join(["\n "] * empty_lines)

            self.multi_cell(width, height, text, border=border, align=align,
                            new_x=new_x, new_y=new_y, wrapmode=WrapMode.CHAR)

    def add_commune_data(self, commune, events):
        self.set_font(FONT_NAME, "B", FONT_SZ_TABLE_HEADER)
        self.row(ROW_HEIGHT_HEADER,
                 ["# Réservation /\nDate & Heure", "Réservateur /\nInstitution",
                  "Titulaire /\n# bon commande", "Activité\n", "\nTarif"],
                 styles=["I", "I", "I", "I", "I"],
                 aligns=["L", "L", "L", "L", "R"])

        total_sum = 0

        self.set_font(FONT_NAME, "", FONT_SZ_TABLE_ROW)
        for entry in events:
            if entry['booking_payment'] != "Facture et bon de commande":
                continue
            
            total_sum += entry['price']

            # Sub-row 1
            self.row(ROW_HEIGHT_ROW, [
                format_booking_number(entry['booking_number']),
                entry['responsable'],
                entry['titulaire'],
                entry['activity'],
                ""
            ], border="", styles=["", "B", "",  "B", ""], aligns=["L", "L", "L", "L", "R"])

            # Sub-row 2
            self.row(ROW_HEIGHT_ROW,
                     [entry['datetime'],
                      entry['customer_name'],
                      entry['bon_commande'],
                      "", # entry['booking_internal_comment'],
                      format_price(entry['price'])],
                     styles=["", "I", "", "", ""],
                     aligns=["L", "L", "L", "L", "R"])

        # Total sum for the commune
        self.set_font(FONT_NAME, "B", FONT_SZ_TABLE_ROW)
        self.row(ROW_HEIGHT_ROW, ["", "", "", "Total", format_price(total_sum)],
                 styles=["", "", "", "B", "B"],
                 aligns=["L", "L", "L", "L", "R"])

def generate_reports(file_path):
    data = defaultdict(list)
    extension = file_path.split(".")[1]
    
    get_col = None # Extracts value from row and col name
    rows = None # Iterates rows
    csvfile = None
    
    if extension == "csv":
        csvfile = open(file_path, newline='')

        def get_col(row, col_name):
            return row[col_name]
        rows = csv.DictReader(csvfile)

    elif extension == 'xlsx':
        wb = load_workbook(filename=file_path)
        sh = wb.active
        max_row = sh.max_row

        row0 = ([r for r in sh.iter_rows(min_row=1, max_row=1)])[0]
        cols = {c.value: c.column for c in row0}

        def get_col(row, col_name):
            v = row[cols[col_name]-1].value
            if type(v) == str:
                return v.strip()
            return v
        rows = sh.iter_rows(min_row=2, max_row=max_row)

    else:
        raise Exception("Cant read this file type. Try csv or xlsx!")

    # Actually read
    for row in rows:
        activity = get_col(row, "Offer\nName")
        customer_invoice_address_name = get_col(row, "Customer\nInvoice address name")
        data[customer_invoice_address_name].append({
            'cia_name': get_col(row, "Customer\nInvoice address name"),
            'cia_street': get_col(row, "Customer\nInvoice address street"),
            'cia_zip': get_col(row, "Customer\nInvoice address postal code"),
            'cia_city': get_col(row, "Customer\nInvoice address city"),
            'activity': activity,
            'responsable': get_col(row, "Booker\nFull name"),
            'customer_name': get_col(row, "Customer\nName"),
            'booking_number': get_col(row, "Booking\nNumber"),
            'booking_payment': get_col(row, "Booking\nPayment"),
            # 'booking_internal_comment': get_col(row, "Booking\nInternal comment"),
            'datetime': get_col(row, "Offer\nStart date & time"),
            'price': float(get_col(row, "Reservation\nPrice")),
            'titulaire': get_col(row, "Property\nNom du titulaire"),
            'bon_commande': get_col(row, "Property\nNuméro bon de commande"),
        })

    if csvfile:
        csvfile.close()

    pdf = PDF()
    
    pdf.set_margins(16, 20, 16) # L T R 
    for customer_invoice_address_name, activities in data.items():
        pdf.add_font(FONT_NAME, '', os.path.join(FONT_PATH, FONT_FILE_REGULAR))
        pdf.add_font(FONT_NAME, 'B', os.path.join(FONT_PATH, FONT_FILE_BOLD))
        pdf.add_font(FONT_NAME, 'I', os.path.join(FONT_PATH, FONT_FILE_OBLIQUE))
        pdf.add_page()
        pdf.new_document()

        # Initial content
        pdf.set_font(FONT_NAME, 'B', FONT_SZ_NORMAL)
        pdf.multi_cell(0, 5, f"Adresse de facturation: ", 0, 'L')
        pdf.ln(2)

        pdf.set_font(FONT_NAME, '', FONT_SZ_NORMAL)
        pdf.set_x(pdf.get_x() + 10)

        a0 = activities[0]
        pdf.multi_cell(0, 5, f"{a0['cia_name']}\n{a0['cia_street']}\n{a0['cia_zip']} {a0['cia_city']}", 0, 'L')
        pdf.ln(10)

        # Adding activity data
        pdf.add_commune_data(customer_invoice_address_name, activities)

    # suffix = datetime.datetime.now().isoformat()
    suffix = datetime.date.today().isoformat()
    filename = f"facture_{suffix}.pdf"
    pdf.output(filename)
    print(f"Generated PDF {filename}")

if __name__ == '__main__':
    import sys
    generate_reports(sys.argv[1])
